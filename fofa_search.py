#!/usr/bin/python3
#-*- coding:utf-8 -*-

import base64
import json
import fofa
import math
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib.request,urllib.error,urllib.parse


def header():
	print('\033[34m*-----------------------------------------*\033[0m')
	print('\033[34m!           Author:Fly_Wuhoo    \033[0m')
	print('\033[34m!  使用格式: python3 fofa_search.py    \033[0m')
	print('\033[34m*-----------------------------------------*\033[0m')


class Client:
	def __init__(self,email,key):
		self.email = email
		self.key = key
		self.base_url = "https://fofa.info"
		self.search_api_url = "/api/v1/search/all"
		self.login_api_url = "/api/v1/info/my"
		self.get_userinfo() #check email and key

	def get_userinfo(self):
		api_full_url = "%s%s" % (self.base_url,self.login_api_url)
		param = {"email":self.email,"key":self.key}
		res = self.__http_get(api_full_url,param)
		return json.loads(res)

	def get_data(self,query_str,page=1,fields=""):
		res = self.get_json_data(query_str,page,fields)
		return json.loads(res)

	def get_json_data(self,query_str,page=1,fields=""):
		api_full_url = "%s%s" % (self.base_url,self.search_api_url)
		param = {"qbase64":base64.b64encode(query_str),"email":self.email,"key":self.key,"page":page,"fields":fields}
		res = self.__http_get(api_full_url,param)
		return res


	def __http_get(self,url,param):
		param = urllib.parse.urlencode(param)
		url = "%s?%s" % (url,param)
		try:
			req = urllib.request.Request(url)
			res = urllib.request.urlopen(req).read()
			if "errmsg" in str(res):
				raise RuntimeError(res)
		except urllib.error.URLError as e:
			print("errmsg："+e.read())
			raise e
		return res


class FoFa:
	#初始化
	def __init__(self):
		self.mail="18850620382@139.com"
		self.key="569459458cb1c02335dbebe8edf162e6"
		try:
			self.api = Client(self.mail,self.key)
			#print('ok')
		except Exception as e:
			print('Error: {}'.format(e))

	def Search(self,dork):
		try:
			result = self.api.get_data(dork.encode())
			pages = math.ceil(result['size']/100)
			print('*-----------------------------------------*')
			print('!        \033[36m共{}页,{}条匹配结果\033[0m'.format(pages,str(result['size'])))
			print('*-----------------------------------------*')
			book = Workbook()
			sheet = book.active
			header=['IP','端口','URL','status','域名','服务','操作系统','国家','title','header']
			for h in range(len(header)):
				sheet.cell(row=1, column=h+1).value = header[h]
			k=2
			with open("Url.txt","a+") as fp:
				for i in range(1,min((math.ceil(result['size']/100)+1),101)):						
					print("正在保存第{}页数据".format(i))
					for ip,port,host,domain,server,os,country,title,header,isp in self.api.get_data(dork.encode(),i,fields="ip,port,host,domain,server,os,country,title,header,isp")['results']:
						try:					
							sheet.cell(row=k, column=1).value = str(ip)
						except Exception as e:
							print(e)
							pass
						try:					
							sheet.cell(row=k, column=2).value = str(port)
						except Exception as e:
							print(e)
							pass
						try:	
							if "https://" in host:				
								sheet.cell(row=k, column=3).value = str(host)
								fp.write(host+"\n")
							else:
								sheet.cell(row=k, column=3).value = str("http://"+host)
								fp.write("http://"+host+"\n")
						except Exception as e:
							print(e)
							pass
						try:
							status = int(header.split(' ')[1].strip())
							if status not in [200, 301, 302, 303, 304, 307, 400, 401, 403, 404, 405, 407,500, 501, 502, 503, 504, 508]:
								status = ''
						except:
							status = ''
						sheet.cell(row=k, column=4).value = status
						try:					
							sheet.cell(row=k, column=5).value = str(domain)
						except Exception as e:
							print(e)
							pass
						try:					
							sheet.cell(row=k, column=6).value = str(server)
						except Exception as e:
							print(e)
							pass					
						try:
							sheet.cell(row=k, column=7).value = str(os)
						except Exception as e:
							print(e)
							pass
						try:
							sheet.cell(row=k, column=8).value = str(country)
						except Exception as e:
							print(e)
							pass
						try:
							sheet.cell(row=k, column=9).value = str(title)
						except Exception as e:
							print(e)
							pass
						try:
							sheet.cell(row=k, column=10).value = str(header)
						except Exception as e:
							print(e)
							pass
						try:
							sheet.cell(row=k, column=11).value = str(isp)
						except Exception as e:
							print(e)
							pass
						k=k+1

						
				try:
					book.save("result.xlsx")
					print('查询结果已保存至当前目录下的result.xlsx!!!\n')
				except Exception as e:
					print('保存文件过程中出现异常，结果保存失败,！！！\n%s'%e)
			fp.close()
					
		except Exception as e:
			print('Error: {}'.format(e))
		

if __name__ == '__main__':
	header()
	dork = str(input("\033[31m请输入Fofa查询语法:\nDork >>> \033[0m")) #定义字符颜色
	F=FoFa()
	F.Search(dork) 
